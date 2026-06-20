import os
import requests
import pandas as pd
import time
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────
# Konfiqurasiya
# ─────────────────────────────────────────────
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36",
    "accept": "application/json, text/plain, */*",
}
TIMEOUT = 30        # hər sorğu üçün maksimum gözləmə (saniyə)
DELAY   = 0.4       # sorğular arasında fasilə (serverə yük verməmək üçün)

# Session + avtomatik retry (HTTP 429/5xx üçün)
session = requests.Session()
retry_strategy = Retry(total=3, backoff_factor=2, status_forcelist=[429, 500, 502, 503, 504])
session.mount("https://", HTTPAdapter(max_retries=retry_strategy))
session.mount("http://",  HTTPAdapter(max_retries=retry_strategy))


def safe_get(url, max_attempts=3):
    """Timeout və şəbəkə xətalarına qarşı davamlı GET."""
    for attempt in range(1, max_attempts + 1):
        try:
            r = session.get(url, headers=HEADERS, timeout=TIMEOUT)
            r.raise_for_status()
            return r
        except requests.exceptions.Timeout:
            print(f"    ⚠  Timeout (cəhd {attempt}/{max_attempts}) — 5s gözlənilir...")
            time.sleep(5)
        except requests.exceptions.RequestException as e:
            print(f"    ⚠  Xəta (cəhd {attempt}/{max_attempts}): {e} — 5s gözlənilir...")
            time.sleep(5)
    print("    ✗  Sorğu uğursuz oldu, keçilir.")
    return None


# ─────────────────────────────────────────────
# 1. Bütün tenderləri topla
# ─────────────────────────────────────────────
PAGE_SIZE_EVENTS = 15
all_events = []
page = 1
total_pages = 1

print("=" * 55)
print("TENDERLƏR TOPLANIR...")
print("=" * 55)

while page <= total_pages:
    url = (
        f"https://etender.gov.az/api/events"
        f"?EventType=2&PageSize={PAGE_SIZE_EVENTS}&PageNumber={page}"
        f"&EventStatus=1&Keyword=&buyerOrganizationName=tikilm%C9%99kd%C9%99%20olan"
        f"&documentNumber=&publishDateFrom=&publishDateTo="
        f"&AwardedparticipantName=&AwardedparticipantVoen=&DocumentViewType="
    )
    r = safe_get(url)
    if r is None:
        print(f"  ✗ Səhifə {page} alına bilmədi, dayandırılır.")
        break

    data = r.json()

    if page == 1:
        total_pages = data["totalPages"]
        total_items = data["totalItems"]
        print(f"  Ümumi tender sayı : {total_items}")
        print(f"  Ümumi səhifə sayı : {total_pages}")
        print("-" * 55)

    all_events.extend(data["items"])
    print(f"  Səhifə {page}/{total_pages} — {len(all_events)} tender toplandı")

    if not data["hasNextPage"]:
        break
    page += 1
    time.sleep(DELAY)

# ─────────────────────────────────────────────
# 2. Tender DataFrame
# ─────────────────────────────────────────────
df_events = pd.DataFrame(all_events)

column_rename = {
    "eventId":                "Tender ID",
    "privateRfxId":           "RFx ID",
    "eventType":              "Tender Növü",
    "eventStatus":            "Status",
    "eventName":              "Tender Adı",
    "buyerOrganizationName":  "Alıcı Təşkilat",
    "publishDate":            "Yayım Tarixi",
    "endDate":                "Bitmə Tarixi",
    "awardedParticipantName": "Qalib İştirakçı",
    "awardedParticipantVoen": "Qalib VÖEN",
    "hasNewVersion":          "Yeni Versiya Var?",
}
df_events = df_events.rename(columns={k: v for k, v in column_rename.items() if k in df_events.columns})

for col in ["Yayım Tarixi", "Bitmə Tarixi"]:
    if col in df_events.columns:
        df_events[col] = pd.to_datetime(df_events[col], errors="coerce").dt.strftime("%d.%m.%Y %H:%M")

event_ids = df_events["Tender ID"].dropna().astype(int).tolist()
print(f"\nCəmi {len(event_ids)} tender ID tapıldı.")

# ─────────────────────────────────────────────
# 3. Hər event ID üçün BOM sətirləri topla
# ─────────────────────────────────────────────
PAGE_SIZE_BOM = 100
all_bom_rows = []

print("\n" + "=" * 55)
print("BOM SƏTİRLƏRİ TOPLANIR...")
print("=" * 55)

for idx, event_id in enumerate(event_ids, 1):
    page = 1
    event_rows = []

    while True:
        url = (
            f"https://etender.gov.az/api/events/{event_id}/bomLines"
            f"?PageSize={PAGE_SIZE_BOM}&PageNumber={page}"
        )
        r = safe_get(url)
        if r is None:
            print(f"  [{idx}/{len(event_ids)}] Event {event_id} — alına bilmədi, keçilir")
            break

        data = r.json()
        items = data.get("items", [])

        if page == 1 and not items:
            break  # Bu tender üçün BOM yoxdur

        for item in items:
            item["Tender ID"] = event_id

        event_rows.extend(items)

        if not data.get("hasNextPage"):
            break
        page += 1
        time.sleep(DELAY)

    all_bom_rows.extend(event_rows)
    print(f"  [{idx}/{len(event_ids)}] Event {event_id} — {len(event_rows)} BOM sətiri")

print(f"\nCəmi BOM sətiri: {len(all_bom_rows)}")

# ─────────────────────────────────────────────
# 4. BOM DataFrame — Tender ID-ni öndə göstər
# ─────────────────────────────────────────────
if all_bom_rows:
    df_bom = pd.DataFrame(all_bom_rows)
    cols = ["Tender ID"] + [c for c in df_bom.columns if c != "Tender ID"]
    df_bom = df_bom[cols]
else:
    df_bom = pd.DataFrame()

# ─────────────────────────────────────────────
# 5. Excel-ə yaz (py faylının olduğu qovluqda)
# ─────────────────────────────────────────────
try:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
except NameError:
    SCRIPT_DIR = os.getcwd()

output_file = os.path.join(SCRIPT_DIR, "etender_tikilmekde_olan_FULL.xlsx")

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALT_FILL    = PatternFill("solid", start_color="DCE6F1")
NORMAL_FONT = Font(name="Arial", size=10)
BD          = Side(style="thin", color="B0B0B0")
CELL_BORDER = Border(left=BD, right=BD, top=BD, bottom=BD)


def style_sheet(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER
    ws.row_dimensions[1].height = 30

    for row_idx in range(2, ws.max_row + 1):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for cell in ws[row_idx]:
            cell.font = NORMAL_FONT
            cell.fill = fill
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="center")

    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 55)

    ws.freeze_panes = "A2"


with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    df_events.to_excel(writer, sheet_name="Tenderlər", index=False)
    if not df_bom.empty:
        df_bom.to_excel(writer, sheet_name="BOM Sətirləri", index=False)
    else:
        writer.book.create_sheet("BOM Sətirləri")

    style_sheet(writer.sheets["Tenderlər"])
    if not df_bom.empty:
        style_sheet(writer.sheets["BOM Sətirləri"])

print(f"\n✅ Fayl hazırdır: {output_file}")
print(f"   • 'Tenderlər'     sheet: {len(df_events)} sətir")
print(f"   • 'BOM Sətirləri' sheet: {len(df_bom)} sətir")